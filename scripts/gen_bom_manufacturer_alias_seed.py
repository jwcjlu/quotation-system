#!/usr/bin/env python3
"""
从「电子元器件常见厂牌和缩写.xlsx」生成 t_bom_manufacturer_alias 的 INSERT SQL。
alias_norm 与 Go NormalizeMfrString 一致：trim + Unicode NFKC + strings.ToUpper（此处用 .upper()）。

合并规则（避免 uk_bom_mfr_alias_norm 冲突）：
- 表中「英文名」简写与全称视为同一厂牌时，统一到规范展示名再生成 canonical_id。
"""
from __future__ import annotations

import argparse
import re
import unicodedata
from collections import defaultdict
from pathlib import Path

import openpyxl

# 英文名 -> 规范展示名（与 canonical 主档一致）
MASTER_ENGLISH: dict[str, str] = {
    "ST": "STMicroelectronics",
    "ON Semi": "ON Semiconductor",
    "Diodes": "Diodes Inc.",
}


def norm_mfr(s: str) -> str:
    s = (s or "").strip()
    if not s:
        return ""
    return unicodedata.normalize("NFKC", s).upper()


def canonical_id(display_name: str) -> str:
    d = (display_name or "").strip()
    slug = re.sub(r"[^a-zA-Z0-9]+", "_", d).strip("_").upper()
    if not slug:
        slug = "UNKNOWN"
    return "MFR_" + slug[:110]


def split_abbrev(s: object) -> list[str]:
    if s is None:
        return []
    s = str(s).strip()
    if not s:
        return []
    parts = re.split(r"[/、，,|]+", s)
    return [p.strip() for p in parts if p.strip()]


def master_display(english: str) -> str:
    e = (english or "").strip()
    return MASTER_ENGLISH.get(e, e)


def collect_rows(xlsx: Path) -> dict[str, tuple[str, str, str]]:
    """alias_norm -> (canonical_id, display_name, alias) 后者取首次出现。"""
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    header = [str(x).strip() if x is not None else "" for x in rows[0]]
    idx_en = header.index("英文名")
    idx_ab = header.index("常用缩写")

    by_norm: dict[str, tuple[str, str, str]] = {}
    conflicts: list[tuple[str, tuple, tuple]] = []

    for row in rows[1:]:
        if not row:
            continue
        raw_en = row[idx_en]
        if raw_en is None or not str(raw_en).strip():
            continue
        english = str(raw_en).strip()
        display = master_display(english)
        cid = canonical_id(display)
        ab_cell = row[idx_ab]

        alias_tokens: list[str] = [english, display]
        alias_tokens.extend(split_abbrev(ab_cell))

        seen_tok = set()
        for al in alias_tokens:
            al = (al or "").strip()
            if not al or al in seen_tok:
                continue
            seen_tok.add(al)
            an = norm_mfr(al)
            if not an:
                continue
            if an in by_norm:
                old_cid, old_disp, old_al = by_norm[an]
                if old_cid != cid:
                    conflicts.append((an, (old_cid, old_disp, old_al), (cid, display, al)))
                continue
            by_norm[an] = (cid, display, al)

    if conflicts:
        msg = "alias_norm 冲突（请补充 MASTER_ENGLISH 或手工删行）：\n"
        for an, a, b in conflicts:
            msg += f"  {an!r}: {a} vs {b}\n"
        raise SystemExit(msg)

    return by_norm


def sql_escape(s: str) -> str:
    return s.replace("\\", "\\\\").replace("'", "''")


def emit_sql(by_norm: dict[str, tuple[str, str, str]], out: Path) -> None:
    lines = [
        "-- 种子数据：来自仓库根目录《电子元器件常见厂牌和缩写.xlsx》",
        "-- 由 scripts/gen_bom_manufacturer_alias_seed.py 生成；请勿手改大段，改 xlsx 后重跑脚本。",
        "-- alias_norm 与 internal/biz/bom_mfr_normalize.go NormalizeMfrString 一致。",
        "",
        "INSERT INTO t_bom_manufacturer_alias (canonical_id, display_name, alias, alias_norm) VALUES",
    ]
    items = sorted(by_norm.items(), key=lambda kv: (kv[1][0], kv[0]))
    value_lines = []
    for an, (cid, disp, al) in items:
        value_lines.append(
            f"  ('{sql_escape(cid)}', '{sql_escape(disp)}', '{sql_escape(al)}', '{sql_escape(an)}')"
        )
    lines.append(",\n".join(value_lines) + ";")
    out.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument(
        "--xlsx",
        type=Path,
        default=Path(__file__).resolve().parent.parent / "电子元器件常见厂牌和缩写.xlsx",
    )
    ap.add_argument(
        "--out",
        type=Path,
        default=Path(__file__).resolve().parent.parent
        / "docs/schema/migrations/20260329_seed_bom_manufacturer_alias_from_xlsx.sql",
    )
    args = ap.parse_args()
    if not args.xlsx.is_file():
        raise SystemExit(f"找不到 xlsx: {args.xlsx}")
    by_norm = collect_rows(args.xlsx)
    emit_sql(by_norm, args.out)
    print(f"wrote {len(by_norm)} rows -> {args.out}")


if __name__ == "__main__":
    main()
